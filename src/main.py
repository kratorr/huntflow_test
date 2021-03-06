import os
import sys
import unicodedata
import logging
import argparse

from mimetypes import MimeTypes

import requests
from requests.exceptions import ConnectionError
from openpyxl import load_workbook

logger = logging.getLogger('huntflow_uploader')
logger.setLevel(logging.DEBUG)
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
ch = logging.StreamHandler()
ch.setFormatter(formatter)
logger.addHandler(ch)

parser = argparse.ArgumentParser(description='Huntflow upload candidates script')
parser.add_argument('-t', '--token', help='API Token', required=True)
parser.add_argument('-f', '--db_file', help='path to db_file', required=True)
args = parser.parse_args()

API_URL = 'https://dev-100-api.huntflow.ru/'

HEADERS = {
    'User-Agent': 'App/1.0 (incaseoffire@example.com)',
    'Authorization': 'Bearer {}'.format(args.token)
}

ENDPOINTS = {
    'vacancies': 'account/{}/vacancies',
    'statuses': 'account/{}/vacancy/statuses',
    'add_applicant': 'account/{}/applicants',
    'upload_cv': 'account/{}/upload',
    'accounts': 'accounts',
    'attach_to_vacancy': 'account/{}/applicants/{}/vacancy'
}

tranlastion_statuses = {
    'Интервью с HR': 'HR Interview',
    'Выставлен оффер': 'Offered',
    'Отправлено письмо': 'Submitted',
    'Отказ': 'Declined'
}


def get_account_id():
    r = requests.get(API_URL + ENDPOINTS['accounts'], headers=HEADERS)

    if r.status_code == requests.codes.ok:
        r = r.json()
    return r['items'][0]['id']


def get_vacancies(account_id):
    url = API_URL + ENDPOINTS['vacancies'].format(account_id)
    try:
        r = requests.get(url, headers=HEADERS)
        if r.status_code == requests.codes.ok:
            r = r.json()
    except requests.exceptions.ConnectionError as e:
        return None
    return r['items']


def get_candidate_statuses(account_id):
    url = API_URL + ENDPOINTS['statuses'].format(account_id)
    try:
        r = requests.get(url, headers=HEADERS)
        if r.status_code == requests.codes.ok:
            r = r.json()
    except requests.exceptions.ConnectionError as e:
        return None
    return r['items']


def load_candidates_from_xls(path):
    """Load candidates from xlsx database."""
    logger.info('Starting loading candidates from xlsx')
    xls_db = load_workbook(filename=path)
    sheet = xls_db.active
    candidates = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        position, fullname, salary, comment, status = row
        candidates.append(
            {
                'position': position,
                'fullname': fullname.strip(),
                'salary': salary,
                'comment': comment,
                'status': status
            }
        )
    logger.info('Сandidates successfully loaded from the xlsx')
    return candidates


def append_cv_file(candidates):
    for canidate in candidates:
        position_path = os.path.abspath(os.path.join('.', canidate['position']))
        for file in os.listdir(position_path):
            if canidate['fullname'] in unicodedata.normalize('NFC', file):  # пришлось нормализовать имя файла
                #  после MacOS некоторые символы были некорректны
                canidate['cv_path'] = os.path.abspath(os.path.join(position_path, file))
    return candidates


def append_extra_data(candidates, vacancies, statuses):
    for candidate in candidates:
        candidate['vacancy_id'] = list(
            filter(lambda x: x['position'] == candidate['position'], vacancies)
        )[0]['id']
        translated_status = tranlastion_statuses[candidate['status']]
        candidate['status_id'] = list(
            filter(lambda x: x['name'] == translated_status, statuses)
        )[0]['id']


def upload_cv(candidate, account_id):
    headres_local = HEADERS.copy()
    headres_local.update({'X-File-Parse': 'true'})
    mime = MimeTypes()
    mime_type = mime.guess_type(candidate['cv_path'])
    files = {'file': (os.path.basename(candidate['cv_path']), open(candidate['cv_path'], 'rb'), mime_type[0])}
    url = API_URL + ENDPOINTS['upload_cv'].format(account_id)
    try:
        r = requests.post(url, files=files, headers=headres_local)
        if r.status_code == requests.codes.ok:
            r = r.json()
    except requests.exceptions.ConnectionError as e:
        print(e)
    return r


def upload_applicant(prepared_candidate, account_id):
    url = API_URL + ENDPOINTS['add_applicant'].format(account_id)
    try:
        r = requests.post(url, headers=HEADERS, json=prepared_candidate)
        if r.status_code == requests.codes.ok:
            r = r.json()
    except ConnectionError as e:
        sys.exit(1)
    return r['id']


def attach_to_vacancy(account_id, applicant_id, attach_params):
    url = API_URL + ENDPOINTS['attach_to_vacancy'].format(account_id, applicant_id)
    payload = {
        "vacancy": attach_params['vacancy_id'],
        "status": attach_params['status_id'],
        "comment": attach_params['comment'],
        "files": [
            {
                "id": attach_params['file_id']
            }
        ],
        "rejection_reason": attach_params['comment'] if attach_params['status_id'] == 50 else None  # 50 Decline status
    }

    try:
        r = requests.post(url, headers=HEADERS, json=payload)
        if r.status_code == requests.codes.ok:
            r = r.json()
    except ConnectionError as e:
        logger.error('Upload error')
        sys.exit(1)


def upload_candidates(candidate, account_id, vacancies, statuses):
    data_from_cv = upload_cv(candidate, account_id)
    prepared_candidate = {
        "last_name": data_from_cv['fields']['name']['last'],
        "first_name": data_from_cv['fields']['name']['first'],
        "middle_name": data_from_cv['fields']['name']['middle'],
        "phone": data_from_cv['fields']['phones'][0],
        "email": data_from_cv['fields']['email'],
        "position": candidate['position'],
        "company": data_from_cv['fields']['experience'][0]['company'],
        "money": str(candidate['salary']),
        "birthday_day": data_from_cv['fields']['birthdate']['day'] \
            if data_from_cv['fields']['birthdate'] is not None else None,
        "birthday_month": data_from_cv['fields']['birthdate']['month'] \
            if data_from_cv['fields']['birthdate'] is not None else None,
        "birthday_year": data_from_cv['fields']['birthdate']['year'] \
            if data_from_cv['fields']['birthdate'] is not None else None,
        "photo": data_from_cv['photo']['id'] if 'photo' in data_from_cv else None,
        "externals": [
            {
                "data": {
                    "body": data_from_cv['text'] if 'text' in data_from_cv else None,
                },
                "auth_type": "NATIVE",
                "files": [
                    {
                        "id": data_from_cv['id']
                    }
                ],
                "account_source": None
            }
        ]
    }
    applicant_id = upload_applicant(prepared_candidate, account_id)
    attach_params = {
        'vacancy_id': candidate['vacancy_id'],
        'status_id': candidate['status_id'],
        'comment': candidate['comment'],
        'file_id': data_from_cv['id']
    }
    attach_to_vacancy(account_id, applicant_id, attach_params)


def main():
    account_id = get_account_id()
    logger.info('Starting load dicts')
    vacancies = get_vacancies(account_id)
    statuses = get_candidate_statuses(account_id)
    if vacancies is None and statuses is None:
        logger.error('Error loading dicts')
        exit(1)
    logger.info('Dicts loaded successfully')
    candidates = load_candidates_from_xls(args.db_file)
    append_cv_file(candidates)
    append_extra_data(candidates, vacancies, statuses)
    logger.info('Starting upload candidates to Huntflow')
    for candidate in candidates:
        upload_candidates(candidate, account_id, vacancies, statuses)
    logger.info('Candidates uploaded successfully')


if __name__ == '__main__':
    main()
